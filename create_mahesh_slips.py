# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import copy, os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from datetime import datetime

template_path = 'C:/Users/Windows 4/Downloads/chethan_salary_slips.xlsx'
output_path = 'C:/Users/Windows 4/Downloads/mahesh_salary_slips.xlsx'

twb = openpyxl.load_workbook(template_path)
tws = twb['Dec 2025']
print('Reading template...')

cell_styles = {}
for row in range(1, 35):
    for col in range(1, 9):
        cell = tws.cell(row=row, column=col)
        cell_styles[(row, col)] = {
            'font': copy.copy(cell.font),
            'fill': copy.copy(cell.fill),
            'border': copy.copy(cell.border),
            'alignment': copy.copy(cell.alignment),
            'number_format': cell.number_format,
        }

merged_ranges = [str(mr) for mr in tws.merged_cells.ranges]
print(f'  {len(merged_ranges)} merged ranges')

col_widths = {}
for cl, dim in tws.column_dimensions.items():
    col_widths[cl] = dim.width
row_heights = {}
for rn, dim in tws.row_dimensions.items():
    row_heights[rn] = dim.height

rupee = chr(8377)

sheets = [
    ('Jan 2025', 'PAY SLIP FOR THE MONTH OF January - 2025', 27, 25, 2, 611.11, 15000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 'Fifteen Thousand Rupees Only', 'January 2025'),
    ('Feb 2025', 'PAY SLIP FOR THE MONTH OF February - 2025', 22, 20, 2, 750.00, 15000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 'Fifteen Thousand Rupees Only', 'February 2025'),
    ('Mar 2025', 'PAY SLIP FOR THE MONTH OF March - 2025', 26, 26, 0, 634.62, 16500, 0, 0, 0, 0, 0, 0, 0, 0, 0, 'Sixteen Thousand Five Hundred Rupees Only', 'March 2025'),
]

wb = Workbook()

for idx, (name, title, td, dp, da, pd_val, basic, hra, conv, spec, ot, pf, esi, pt, tds, adv, words, month_yr) in enumerate(sheets):
    if idx == 0:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(title=name)
    print(f'Creating: {name}')

    for mr in merged_ranges:
        ws.merge_cells(mr)
    for cl, w in col_widths.items():
        ws.column_dimensions[cl].width = w
    for rn, h in row_heights.items():
        ws.row_dimensions[rn].height = h

    for (row, col), style in cell_styles.items():
        cell = ws.cell(row=row, column=col)
        cell.font = copy.copy(style['font'])
        cell.fill = copy.copy(style['fill'])
        cell.border = copy.copy(style['border'])
        cell.alignment = copy.copy(style['alignment'])
        cell.number_format = style['number_format']

    ws['A1'] = 'PRINTIGLY'
    ws['A2'] = 'Digital Printing & Corporate Gifting Solutions | Since 1984'
    ws['A3'] = 'Door No 176, 1st Main, 19th Cross, Agrahara Layout, Yelahanka, Bengaluru, Karnataka - 560064'
    ws['A4'] = 'Phone: 7204910047 | GSTIN: 29ABBFP4851E1Z7'
    ws['A6'] = title
    ws['A8'] = 'Employee Code:'
    ws['B8'] = 202404
    ws['E8'] = 'Department:'
    ws['F8'] = 'Digital Marketing'
    ws['A9'] = 'Employee Name:'
    ws['B9'] = 'Poluru Mahesh'
    ws['E9'] = 'Designation:'
    ws['F9'] = 'Digital Marketing'
    ws['A10'] = 'Date of Joining:'
    ws['B10'] = datetime(2024, 11, 11)
    ws['E10'] = 'Bank A/C No:'
    ws['F10'] = 3265314429
    ws['A11'] = 'PAN No:'
    ws['B11'] = 'DESPP0227P'
    ws['E11'] = 'UAN No:'
    ws['F11'] = '-'
    ws['A13'] = 'ATTENDANCE DETAILS'
    ws['A14'] = 'Total Days'
    ws['B14'] = td
    ws['C14'] = 'Days Present'
    ws['D14'] = dp
    ws['E14'] = 'Days Absent'
    ws['F14'] = da
    ws['G14'] = f'Per Day ({rupee})'
    ws['H14'] = pd_val
    ws['A16'] = 'EARNINGS'
    ws['E16'] = 'DEDUCTIONS'
    ws['A17'] = 'Component'
    ws['B17'] = f'Amount ({rupee})'
    ws['E17'] = 'Component'
    ws['F17'] = f'Amount ({rupee})'
    ws['A18'] = 'Basic Salary'
    ws['B18'] = basic
    ws['E18'] = 'Provident Fund (PF)'
    ws['F18'] = pf
    ws['A19'] = 'HRA'
    ws['B19'] = hra
    ws['E19'] = 'ESI'
    ws['F19'] = esi
    ws['A20'] = 'Conveyance'
    ws['B20'] = conv
    ws['E20'] = 'Professional Tax'
    ws['F20'] = pt
    ws['A21'] = 'Special Allowance'
    ws['B21'] = spec
    ws['E21'] = 'TDS'
    ws['F21'] = tds
    ws['A22'] = 'Overtime'
    ws['B22'] = ot
    ws['E22'] = 'Advance'
    ws['F22'] = adv
    ws['A23'] = 'TOTAL EARNINGS'
    ws['B23'] = '=SUM(B18:B22)'
    ws['E23'] = 'TOTAL DEDUCTIONS'
    ws['F23'] = '=SUM(F18:F22)'
    ws['A25'] = 'NET PAYABLE'
    ws['E25'] = '=B23-F23'
    ws['A26'] = f'Amount in Words: {words}'
    ws['A28'] = f'Salary for the Month of {month_yr} has been credited to your HDFC Bank | Branch: Empire Infantry Road'
    ws['A30'] = '* This is a computer-generated pay slip and does not require a signature.'
    ws['A32'] = 'Employee Signature'
    ws['F32'] = 'For Printigly'
    ws['F33'] = 'Authorized Signatory'

wb.save(output_path)
print(f'Saved: {output_path}')

vwb = openpyxl.load_workbook(output_path)
print(f'Sheets: {vwb.sheetnames}')
for sn in vwb.sheetnames:
    vws = vwb[sn]
    mc = len(list(vws.merged_cells.ranges))
    print(f'  {sn}: {mc} merges, A6={vws["A6"].value}')
    print(f'    B18={vws["B18"].value}, E25={vws["E25"].value}, E25_fmt={vws["E25"].number_format}')
    print(f'    A6 font={vws["A6"].font.name}/{vws["A6"].font.size}/{vws["A6"].font.bold}')
    print(f'    A6 fill={vws["A6"].fill.patternType}/{vws["A6"].fill.fgColor.rgb}')
    print(f'    A25 fill={vws["A25"].fill.patternType}/{vws["A25"].fill.fgColor.rgb}')
    print(f'    E25 fill={vws["E25"].fill.patternType}/{vws["E25"].fill.fgColor.rgb}')
    print(f'    A17 border.left={vws["A17"].border.left.style}')
    print(f'    B18 nf={vws["B18"].number_format}')

fsize = os.path.getsize(output_path)
print(f'File size: {fsize:,} bytes')
print('SUCCESS')
