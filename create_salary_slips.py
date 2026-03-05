import openpyxl, copy, datetime, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
wb = openpyxl.Workbook()
months_data = [
    {
        "sheet_name": "Feb 2025",
        "title": "PAY SLIP FOR THE MONTH OF February - 2025",
        "total_days": 22, "days_present": 22, "days_absent": 0,
        "per_day": round(15000 / 22, 2), "basic_salary": 15000.00,
        "hra": 0, "conveyance": 0, "special_allowance": 0, "overtime": 0,
        "pf": 0, "esi": 0, "prof_tax": 0, "tds": 0, "advance": 0,
        "amount_words": "Fifteen Thousand Rupees Only",
        "bank_credit": "Salary Credited to: HDFC Bank | Branch: Empire Infantry Road",
    },
    {
        "sheet_name": "Mar 2025",
        "title": "PAY SLIP FOR THE MONTH OF March - 2025",
        "total_days": 26, "days_present": 26, "days_absent": 0,
        "per_day": round(16500 / 26, 2), "basic_salary": 16500.00,
        "hra": 0, "conveyance": 0, "special_allowance": 0, "overtime": 0,
        "pf": 0, "esi": 0, "prof_tax": 0, "tds": 0, "advance": 0,
        "amount_words": "Sixteen Thousand Five Hundred Rupees Only",
        "bank_credit": "Salary Credited to: HDFC Bank | Branch: Empire Infantry Road",
    },
    {
        "sheet_name": "Apr 2025",
        "title": "PAY SLIP FOR THE MONTH OF April - 2025",
        "total_days": 26, "days_present": 26, "days_absent": 0,
        "per_day": round(16500 / 26, 2), "basic_salary": 16500.00,
        "hra": 0, "conveyance": 0, "special_allowance": 0, "overtime": 0,
        "pf": 0, "esi": 0, "prof_tax": 0, "tds": 0, "advance": 0,
        "amount_words": "Sixteen Thousand Five Hundred Rupees Only",
        "bank_credit": "Salary Credited to: HDFC Bank | Branch: Empire Infantry Road",
    },
]

font_title = Font(name="Cambria", size=20, bold=True, color=Color(rgb="FF1F4E79"))
font_subtitle = Font(name="Cambria", size=10, bold=False, color=Color(rgb="FF555555"))
font_address = Font(name="Cambria", size=9, bold=False)
font_phone = Font(name="Cambria", size=9, bold=False, color=Color(rgb="FF444444"))
font_payslip_title = Font(name="Cambria", size=14, bold=True, color=Color(rgb="FFFFFFFF"))
font_label = Font(name="Cambria", size=10, bold=True)
font_value = Font(name="Calibri", size=11, bold=False, color=Color(theme=1))
font_section_header = Font(name="Cambria", size=11, bold=True)
font_total_label = Font(name="Cambria", size=10, bold=True)
font_net_payable = Font(name="Cambria", size=12, bold=True)
font_amount_words = Font(name="Cambria", size=10, bold=False)
font_bank_credit = Font(name="Cambria", size=9, bold=False, color=Color(rgb="FF444444"))
font_disclaimer = Font(name="Cambria", size=9, bold=False, color=Color(rgb="FF666666"))
font_signature = Font(name="Cambria", size=9, bold=False)
font_for_company = Font(name="Cambria", size=9, bold=True)
font_auth_signatory = Font(name="Cambria", size=9, bold=False)

fill_blue_header = PatternFill(patternType="solid", fgColor="FF1F4E79")
fill_grey = PatternFill(patternType="solid", fgColor="FFD6DCE5")
fill_green = PatternFill(patternType="solid", fgColor="FFC6EFCE")

align_center = Alignment(horizontal="center")

thin_side = Side(style="thin", color=Color(auto=True))
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_no_right = Border(left=thin_side, top=thin_side, bottom=thin_side)

rupee = chr(8377)

for idx, data in enumerate(months_data):
    if idx == 0:
        ws = wb.active
        ws.title = data["sheet_name"]
    else:
        ws = wb.create_sheet(title=data["sheet_name"])

    ws.column_dimensions["A"].width = 15.0
    ws.column_dimensions["B"].width = 18.0
    ws.column_dimensions["C"].width = 15.0

    ws.row_dimensions[1].height = 25.5
    ws.row_dimensions[6].height = 18.0
    ws.row_dimensions[25].height = 15.75

    ws.merge_cells("A1:H1")
    ws["A1"].value = "PRINTIGLY"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:H2")
    ws["A2"].value = "Digital Printing & Corporate Gifting Solutions | Since 1984"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_center

    ws.merge_cells("A3:H3")
    ws["A3"].value = "Door No 176, 1st Main, 19th Cross, Agrahara Layout, Yelahanka, Bengaluru, Karnataka - 560064"
    ws["A3"].font = font_address
    ws["A3"].alignment = align_center

    ws.merge_cells("A4:H4")
    ws["A4"].value = "Phone: 7204910047 | GSTIN: 29ABBFP4851E1Z7"
    ws["A4"].font = font_phone
    ws["A4"].alignment = align_center

    ws.merge_cells("A6:H6")
    ws["A6"].value = data["title"]
    ws["A6"].font = font_payslip_title
    ws["A6"].fill = fill_blue_header
    ws["A6"].alignment = align_center

    ws["A8"].value = "Employee Code:"
    ws["A8"].font = font_label
    ws["B8"].value = 202404
    ws["B8"].font = font_value
    ws["E8"].value = "Department:"
    ws["E8"].font = font_label
    ws.merge_cells("F8:H8")
    ws["F8"].value = "Digital Marketing"
    ws["F8"].font = font_value

    ws["A9"].value = "Employee Name:"
    ws["A9"].font = font_label
    ws["B9"].value = "Poluru Mahesh "
    ws["B9"].font = font_value
    ws["E9"].value = "Designation:"
    ws["E9"].font = font_label
    ws.merge_cells("F9:H9")
    ws["F9"].value = "Digital Marketing"
    ws["F9"].font = font_value

    ws["A10"].value = "Date of Joining:"
    ws["A10"].font = font_label
    ws["B10"].value = datetime.datetime(2024, 11, 11)
    ws["B10"].font = font_value
    ws["B10"].number_format = "mm-dd-yy"
    ws["E10"].value = "Bank A/C No:"
    ws["E10"].font = font_label
    ws.merge_cells("F10:H10")
    ws["F10"].value = 3265314429
    ws["F10"].font = font_value

    ws["A11"].value = "PAN No:"
    ws["A11"].font = font_label
    ws["B11"].value = "DESPP0227P"
    ws["B11"].font = font_value
    ws["E11"].value = "UAN No:"
    ws["E11"].font = font_label
    ws.merge_cells("F11:H11")
    ws["F11"].value = "-"
    ws["F11"].font = font_value

    ws.merge_cells("A13:H13")
    ws["A13"].value = "ATTENDANCE DETAILS"
    ws["A13"].font = font_section_header
    ws["A13"].fill = fill_grey
    ws["A13"].alignment = align_center

    ws["A14"].value = "Total Days"
    ws["A14"].font = font_label
    ws["B14"].value = data["total_days"]
    ws["B14"].font = font_value
    ws["C14"].value = "Days Present"
    ws["C14"].font = font_label
    ws["D14"].value = data["days_present"]
    ws["D14"].font = font_value
    ws["E14"].value = "Days Absent"
    ws["E14"].font = font_label
    ws["F14"].value = data["days_absent"]
    ws["F14"].font = font_value
    ws["G14"].value = "Per Day (" + rupee + ")"
    ws["G14"].font = font_label
    ws["H14"].value = data["per_day"]
    ws["H14"].font = font_value

    ws.merge_cells("A16:D16")
    ws["A16"].value = "EARNINGS"
    ws["A16"].font = font_section_header
    ws["A16"].fill = fill_grey
    ws["A16"].alignment = align_center
    ws.merge_cells("E16:H16")
    ws["E16"].value = "DEDUCTIONS"
    ws["E16"].font = font_section_header
    ws["E16"].fill = fill_grey
    ws["E16"].alignment = align_center

    ws["A17"].value = "Component"
    ws["A17"].font = font_label
    ws["A17"].border = border_all
    ws.merge_cells("B17:D17")
    ws["B17"].value = "Amount (" + rupee + ")"
    ws["B17"].font = font_label
    ws["B17"].border = border_all
    ws["E17"].value = "Component"
    ws["E17"].font = font_label
    ws["E17"].border = border_all
    ws.merge_cells("F17:H17")
    ws["F17"].value = "Amount (" + rupee + ")"
    ws["F17"].font = font_label
    ws["F17"].border = border_all

    earnings = [
        ("Basic Salary", data["basic_salary"]),
        ("HRA", data["hra"]),
        ("Conveyance", data["conveyance"]),
        ("Special Allowance", data["special_allowance"]),
        ("Overtime", data["overtime"]),
    ]
    deductions = [
        ("Provident Fund (PF)", data["pf"]),
        ("ESI", data["esi"]),
        ("Professional Tax", data["prof_tax"]),
        ("TDS", data["tds"]),
        ("Advance", data["advance"]),
    ]

    for i in range(5):
        e_name, e_val = earnings[i]
        d_name, d_val = deductions[i]
        row = 18 + i
        ws.cell(row=row, column=1, value=e_name).font = font_value
        ws.cell(row=row, column=1).border = border_all
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=e_val).font = font_value
        ws.cell(row=row, column=2).border = border_no_right
        ws.cell(row=row, column=2).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=d_name).font = font_value
        ws.cell(row=row, column=5).border = border_all
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        ws.cell(row=row, column=6, value=d_val).font = font_value
        ws.cell(row=row, column=6).border = border_no_right
        ws.cell(row=row, column=6).number_format = "#,##0.00"

    ws["A23"].value = "TOTAL EARNINGS"
    ws["A23"].font = font_total_label
    ws["A23"].fill = fill_grey
    ws["A23"].border = border_all
    ws.merge_cells("B23:D23")
    ws["B23"].value = "=SUM(B18:B22)"
    ws["B23"].font = font_value
    ws["B23"].fill = fill_grey
    ws["B23"].border = border_no_right
    ws["B23"].number_format = "#,##0.00"

    ws["E23"].value = "TOTAL DEDUCTIONS"
    ws["E23"].font = font_total_label
    ws["E23"].fill = fill_grey
    ws["E23"].border = border_all
    ws.merge_cells("F23:H23")
    ws["F23"].value = "=SUM(F18:F22)"
    ws["F23"].font = font_value
    ws["F23"].fill = fill_grey
    ws["F23"].border = border_no_right
    ws["F23"].number_format = "#,##0.00"

    ws.merge_cells("A25:D25")
    ws["A25"].value = "NET PAYABLE"
    ws["A25"].font = font_net_payable
    ws["A25"].fill = fill_green
    ws["A25"].border = Border(left=thin_side, top=thin_side, bottom=thin_side)

    ws.merge_cells("E25:H25")
    ws["E25"].value = "=B23-F23"
    ws["E25"].font = font_net_payable
    ws["E25"].fill = fill_green
    ws["E25"].border = Border(left=thin_side, top=thin_side, bottom=thin_side)
    ws["E25"].number_format = chr(92) + rupee + "#,##0.00"

    ws.merge_cells("A26:H26")
    ws["A26"].value = "Amount in Words: " + data["amount_words"]
    ws["A26"].font = font_amount_words

    ws.merge_cells("A28:H28")
    ws["A28"].value = data["bank_credit"]
    ws["A28"].font = font_bank_credit

    ws.merge_cells("A30:H30")
    ws["A30"].value = "* This is a computer-generated pay slip and does not require a signature."
    ws["A30"].font = font_disclaimer
    ws["A30"].alignment = align_center

    ws["A32"].value = "Employee Signature"
    ws["A32"].font = font_signature
    ws.merge_cells("F32:H32")
    ws["F32"].value = "For Printigly"
    ws["F32"].font = font_for_company

    ws.merge_cells("F33:H33")
    ws["F33"].value = "Authorized Signatory"
    ws["F33"].font = font_auth_signatory

    print("  Sheet created: " + data["sheet_name"])

import os
output_path = os.path.join(os.path.expanduser("~"), "Downloads", "mahesh_salary_slips.xlsx")
wb.save(output_path)
print("File saved successfully to: " + output_path)
print("Sheets: " + str(wb.sheetnames))
